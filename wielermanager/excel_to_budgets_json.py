import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any


def slugify_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_value.lower().strip()
    lowered = lowered.replace("'", "")
    lowered = re.sub(r"[^a-z0-9]+", "-", lowered)
    lowered = re.sub(r"-+", "-", lowered).strip("-")
    return lowered


def parse_cost(value: Any) -> int | float | None:
    if value is None:
        return None

    if isinstance(value, str):
        cleaned = value.strip().replace(",", ".")
        if cleaned == "":
            return None
        try:
            numeric = float(cleaned)
        except ValueError:
            return None
    elif isinstance(value, (int, float)):
        numeric = float(value)
    else:
        return None

    if numeric.is_integer():
        return int(numeric)
    return numeric


def read_budgets_from_excel(input_path: Path, sheet_name: str | None, slugify: bool) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install it with: pip install openpyxl"
        ) from error

    workbook = load_workbook(input_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    budgets: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=1):
        name_cell = row[0] if len(row) > 0 else None
        cost_cell = row[2] if len(row) > 2 else None

        raw_name = name_cell.value if name_cell else None
        raw_cost = cost_cell.value if cost_cell else None

        if raw_name is None or raw_cost is None:
            continue

        name = str(raw_name).strip()
        if not name:
            continue

        cost = parse_cost(raw_cost)
        if cost is None:
            continue

        rider_name = slugify_name(name) if slugify else name
        budgets.append({"name": rider_name, "cost": cost})

    return budgets


def convert_excel_to_json(input_file: Path, output_file: Path, sheet_name: str | None, slugify: bool) -> None:
    budgets = read_budgets_from_excel(input_file, sheet_name=sheet_name, slugify=slugify)
    payload = {"budgets": budgets}

    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text(json.dumps(payload, ensure_ascii=False, indent=4), encoding="utf-8")

    print(f"Wrote {len(budgets)} riders to {output_file}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert cyclists-giro-m-26.xlsx into WIELERMANAGER budgets JSON format."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("cyclists-giro-m-26.xlsx"),
        help="Path to the input .xlsx file (default: cyclists-giro-m-26.xlsx)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("wielermanager/WIELERMANAGER_BUDGETS_giro_2026.json"),
        help="Path to the output JSON file",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Optional sheet name. If omitted, the active sheet is used.",
    )
    parser.add_argument(
        "--no-slugify",
        action="store_true",
        help="Keep names exactly as in Excel (default behavior is slugified names).",
    )

    args = parser.parse_args()

    convert_excel_to_json(
        input_file=args.input,
        output_file=args.output,
        sheet_name=args.sheet,
        slugify=not args.no_slugify,
    )


if __name__ == "__main__":
    main()
